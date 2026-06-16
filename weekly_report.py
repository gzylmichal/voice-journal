#!/usr/bin/env python3
"""
weekly_report.py — Weekly gym coaching report + journal digest + Excel export.

Fetches the last N weeks of workout data from the Notion Workout Log DB and
daily journal entries from the Notion Journal DB, generates a combined AI
analysis, posts it to a Notion page, and optionally exports all workout
history to a fixed Excel file.

Usage:
    python3 weekly_report.py                      # analyse last 4 weeks
    python3 weekly_report.py --weeks 8            # analyse last 8 weeks
    python3 weekly_report.py --dry-run            # print report, don't write to Notion
    python3 weekly_report.py --excel              # also export workout history to Excel
    python3 weekly_report.py --excel-all          # export full all-time history to Excel

Scheduled via systemd timer: Sunday 05:30 — replaces Morning Debrief on Sundays.

Requires in .env:
    NOTION_TOKEN
    NOTION_WORKOUT_DB_ID
    NOTION_TRAINER_PAGE_ID    (blank Notion page — report is appended each week)
    NOTION_DATABASE_ID        (journal DB — for weekly notes context)
    ANTHROPIC_API_KEY or GOOGLE_API_KEY
"""

import os
import sys
import re
import json
import logging
import argparse
import math
import smtplib
import ssl
import time
import requests
import ai_client
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from email.utils import formatdate, make_msgid
from pathlib import Path
from datetime import date, datetime, timedelta
from typing import Optional
from collections import defaultdict
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

NOTION_TOKEN           = os.getenv("NOTION_TOKEN", "")
NOTION_WORKOUT_DB_ID   = os.getenv("NOTION_WORKOUT_DB_ID", "")
NOTION_TRAINER_PAGE_ID = os.getenv("NOTION_TRAINER_PAGE_ID", "")
NOTION_DATABASE_ID     = os.getenv("NOTION_DATABASE_ID", "")   # journal DB
ANTHROPIC_API_KEY      = os.getenv("ANTHROPIC_API_KEY", "")
GOOGLE_API_KEY         = os.getenv("GOOGLE_API_KEY", "")
CLAUDE_MODEL           = os.getenv("CLAUDE_MODEL", "claude-sonnet-4-20250514")
GEMINI_MODEL           = os.getenv("GEMINI_MODEL", "gemini-2.5-flash")
NOTION_VERSION         = "2022-06-28"

# SMTP — reuses same config as Morning Debrief
SMTP_HOST     = os.getenv("SMTP_HOST", "")
SMTP_PORT     = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER     = os.getenv("SMTP_USER", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
EMAIL_FROM    = os.getenv("EMAIL_FROM", "")
EMAIL_TO      = os.getenv("EMAIL_TO", "")

# Fixed path — same file overwritten each Sunday with full history
EXCEL_PATH = Path(os.getenv("EXCEL_PATH", "/opt/voice-journal/reports/workouts.xlsx"))

LOG_FILE = Path(__file__).parent / "weekly_report.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.FileHandler(LOG_FILE), logging.StreamHandler()],
)
log = logging.getLogger("weekly_report")

try:
    from analytics import compute_metrics, epley_1rm, identify_benchmark, score_adherence
    _ANALYTICS_AVAILABLE = True
except ImportError:
    _ANALYTICS_AVAILABLE = False

try:
    from pipeline.plan_config import load_plan_config
    _PLAN_CONFIG_AVAILABLE = True
except ImportError:
    _PLAN_CONFIG_AVAILABLE = False

from models import parse_workout_entry
from pipeline.notion_client import fetch_bodyweight_entries, fetch_metrics_entries


# ---------------------------------------------------------------------------
# Combined system prompt — workout coaching + weekly journal digest
# ---------------------------------------------------------------------------

WEEKLY_REPORT_PROMPT = """You are a strength coach analysing pre-computed training analytics for a specific athlete.

## Athlete Profile
- Primary goal: STRENGTH. All programming decisions prioritise strength on the big three.
- Three benchmark lifts: Bench Press, Deadlift, Squat.
- Pyramid protocol per benchmark lift (working sets only — warm-ups not logged):
    5 × 80% 1RM
    3 × 85% 1RM
    1 × 90% 1RM  (top set — basis for estimated 1RM)
    3 × 80% 1RM  (back-off)
    5 × 70–75% 1RM (back-off)
- Arms and accessories are supporting work, not primary focus.
- Legs intentionally undertrained — squat is the only lower-body benchmark.
- Memos may be in English or Polish — treat both as the same athlete.

## Input
You are receiving PRE-COMPUTED ANALYTICS — not raw logs.
The metrics are ground truth. Do not re-derive, re-calculate, or second-guess them.
Focus entirely on interpretation and prescription.

## Output
Produce a coaching report with EXACTLY these sections, in order:

### 1. Weekly Snapshot
2–3 sentences. Honest overall assessment. If it was a poor week, say so directly.

### 2. Lift Analysis
For EACH of Bench / Deadlift / Squat, write the lift name as **bold** (e.g. **Bench Press**) — NOT as a ### heading:
- 1–2 sentences of progress assessment (reference the trend and plateau_risk from the analytics).
- MANDATORY next-session pyramid — ALL FIVE TIERS, exact weights rounded to nearest 2.5 kg:
  Next session: 5×Xkg → 3×Xkg → 1×Xkg → 3×Xkg → 5×Xkg
- If plateau_risk = high: prescribe a SPECIFIC intervention (e.g. "drop top set to Xkg, rebuild
  over 3 weeks with 5×3 protocol") — not generic advice.
- If insufficient_data for a lift: note it and skip the prescription for that lift.

### 3. Accountability
State sessions logged vs expected (3 per week). Name any benchmark lift missed in any week.
If fully on track, acknowledge in one sentence and move on.

### 4. Supporting Work
Assess push/pull balance from the analytics. Flag any undertrained muscle group that directly
limits a benchmark lift — state the consequence (e.g. "weak triceps limits bench lockout at
top of the rep"). Skip if nothing to flag.

### 5. Recovery Guidance
Include ONLY if ANY of these signals are present in the analytics:
- plateau_risk = high on 2 or more lifts
- adherence_score < 80
- back-off sets absent in recent sessions for any lift
If present: prescribe specific actions (e.g. "deload this week — all weights -15%, 3×5 only,
no top sets"). If no signals present: OMIT this section entirely.

### 6. This Week's 3 Priorities
Numbered list. Ordered by impact on strength outcomes. Specific and actionable.

## Tone
Direct. Evidence-based. No filler, no hedging, no excessive praise.
Prescribe specific numbers — "102.5 kg", not "around 100 kg".
Motivate when performance genuinely warrants it. Call out missed sessions by name.
Output clean markdown only. Use ### ONLY for the 6 top-level section headers. Within sections, use **bold** for lift names and emphasis."""


# ---------------------------------------------------------------------------
# Fetch workout data from Notion
# ---------------------------------------------------------------------------

def fetch_workout_entries(weeks: int) -> list[dict]:
    since = (date.today() - timedelta(weeks=weeks)).isoformat()

    headers = {
        "Authorization": f"Bearer {NOTION_TOKEN}",
        "Content-Type": "application/json",
        "Notion-Version": NOTION_VERSION,
    }

    payload = {
        "filter": {
            "property": "Date",
            "date": {"on_or_after": since},
        },
        "sorts": [{"property": "Date", "direction": "ascending"}],
        "page_size": 200,
    }

    entries = []
    url = f"https://api.notion.com/v1/databases/{NOTION_WORKOUT_DB_ID}/query"

    while True:
        resp = requests.post(url, headers=headers, json=payload, timeout=20)
        resp.raise_for_status()
        data = resp.json()
        entries.extend(data.get("results", []))
        if data.get("has_more"):
            payload["start_cursor"] = data["next_cursor"]
        else:
            break

    log.info(f"Fetched {len(entries)} workout entries from last {weeks} weeks")
    return entries


def fetch_all_workout_entries() -> list[dict]:
    """Fetch every workout entry from Notion (no date filter) for all-time graphs."""
    headers = {
        "Authorization": f"Bearer {NOTION_TOKEN}",
        "Content-Type": "application/json",
        "Notion-Version": NOTION_VERSION,
    }
    payload = {
        "sorts": [{"property": "Date", "direction": "ascending"}],
        "page_size": 200,
    }
    entries = []
    url = f"https://api.notion.com/v1/databases/{NOTION_WORKOUT_DB_ID}/query"
    while True:
        resp = requests.post(url, headers=headers, json=payload, timeout=20)
        resp.raise_for_status()
        data = resp.json()
        entries.extend(data.get("results", []))
        if data.get("has_more"):
            payload["start_cursor"] = data["next_cursor"]
        else:
            break
    log.info(f"Fetched {len(entries)} all-time workout entries for progression graph")
    return entries


def parse_entry(page: dict) -> dict:
    return parse_workout_entry(page)


# ---------------------------------------------------------------------------
# Format workout data for AI prompt
# ---------------------------------------------------------------------------

def format_metrics_for_llm(metrics: dict) -> str:
    """Serialise compute_metrics() output to compact structured text for the LLM."""
    lines = []

    aw = metrics.get("analysis_window", {})
    lines.append(
        f"=== COMPUTED ANALYTICS — {aw.get('start','?')} to {aw.get('end','?')} "
        f"({aw.get('weeks','?')} weeks) ==="
    )
    lines.append("")

    lines.append("--- Strength Progression ---")
    for lift in ("bench", "deadlift", "squat"):
        data = metrics["strength_progression"].get(lift, {})
        if data.get("insufficient_data"):
            lines.append(f"{lift}: insufficient data (< 2 sessions in window)")
            lines.append("")
            continue

        lines.append(
            f"{lift}: e1rm={data['current_e1rm']}kg, trend={data['trend']}, "
            f"velocity={data['velocity_kg_per_week']:+.1f}kg/week, "
            f"plateau_risk={data['plateau_risk']}"
        )
        sessions = data.get("recent_sessions", [])
        if sessions:
            last = sessions[-1]
            pyr = last.get("pyramid", [])
            pyr_str = " | ".join(
                f"{r}×{w:.1f}kg" for w, r in pyr if w > 0 and r > 0
            ) or "—"
            back = "✓" if last["back_off_present"] else "✗"
            lines.append(
                f"  last session ({last['date']}): {pyr_str} · "
                f"vol={last['volume_load_kg']:.0f}kg · back-off {back}"
            )
            if len(sessions) >= 2:
                prior = " | ".join(
                    f"{s['date']} → top={s['top_set']['weight_kg']:.1f}kg×{s['top_set']['reps']}"
                    for s in sessions[:-1]
                )
                lines.append(f"  prior: {prior}")
        lines.append("")

    lines.append("--- Adherence ---")
    adh = metrics.get("adherence", {})
    if adh:
        lines.append(
            f"score={adh['adherence_score']}/100 · "
            f"{adh['total_sessions']}/{adh['expected_sessions']} sessions · "
            f"{adh['avg_sessions_per_week']}/week · streak={adh['current_streak_weeks']} weeks"
        )
        week_parts = []
        for wb in adh.get("weekly_breakdown", []):
            b  = "✓" if wb["bench"]    else "✗"
            dl = "✓" if wb["deadlift"] else "✗"
            sq = "✓" if wb["squat"]    else "✗"
            flag = "⚠" if wb["sessions"] < 3 else ""
            week_parts.append(
                f"{wb['week']}: bench{b} deadlift{dl} squat{sq} ({wb['sessions']}{flag})"
            )
        lines.append(" | ".join(week_parts))
        bf = adh.get("benchmark_frequency", {})
        lines.append(
            f"benchmark frequency: bench={bf.get('bench',0):.2f}/wk · "
            f"deadlift={bf.get('deadlift',0):.2f}/wk · squat={bf.get('squat',0):.2f}/wk"
        )
    lines.append("")

    lines.append("--- Volume Distribution ---")
    vol = metrics.get("volume_distribution", {})
    if vol:
        ratio = vol.get("push_pull_ratio")
        if ratio is not None:
            flag = " (push-heavy)" if ratio > 1.2 else (" (pull-heavy)" if ratio < 0.8 else " (balanced)")
            lines.append(f"push_pull_ratio={ratio:.1f}{flag}")
        else:
            lines.append("push_pull_ratio=N/A (no pull data)")

        mg_parts = " · ".join(
            f"{mg}={avg:.1f}"
            for mg, avg in sorted(
                vol.get("avg_sets_per_week", {}).items(), key=lambda x: -x[1]
            )
        )
        if mg_parts:
            lines.append(f"sets/week: {mg_parts}")

        ut = vol.get("undertrained_groups", [])
        if ut:
            ut_parts = ", ".join(
                f"{mg} ({vol['avg_sets_per_week'].get(mg, 0):.1f}/week)" for mg in ut
            )
            lines.append(f"undertrained: {ut_parts}")

    rpe = metrics.get("rpe_signals", {})
    avg_rpe = rpe.get("avg_rpe_by_exercise", {})
    fatigue_flags = rpe.get("fatigue_flags", [])
    pain_patterns = rpe.get("pain_patterns", [])

    if avg_rpe or fatigue_flags or pain_patterns:
        lines.append("")
        lines.append("--- RPE & Fatigue Signals ---")

        if avg_rpe:
            for ex, week_data in sorted(avg_rpe.items()):
                weekly_str = " | ".join(f"{w}: {r}" for w, r in sorted(week_data.items()))
                lines.append(f"  {ex} avg RPE: {weekly_str}")

        for flag in fatigue_flags:
            lines.append(
                f"  ⚠ FATIGUE FLAG — {flag['lift']}: {flag['reason']}"
            )

        for pp in pain_patterns:
            lines.append(
                f"  ⚠ PAIN PATTERN — {pp['body_part']}: mentioned {pp['occurrences']}× in window"
            )

    # --- Plateau Alerts ---
    fatigue_flag_lifts = {f["lift"] for f in fatigue_flags}
    plateau_lines = []
    for lift in ("bench", "deadlift", "squat"):
        sp = metrics.get("strength_progression", {}).get(lift, {})
        weekly_e1rm = sp.get("weekly_e1rm")
        if not weekly_e1rm or len(weekly_e1rm) < 3:
            continue
        last3 = weekly_e1rm[-3:]
        w1_val = last3[0][1]
        w3_val = last3[2][1]
        if w3_val > w1_val:
            # Rising — no flag
            continue
        vals_str = " → ".join(f"{v:.1f}" for _, v in last3) + " kg"
        has_fatigue = lift in fatigue_flag_lifts
        if has_fatigue:
            suggestion = "deload recommended: −15% all weights this week, no top sets"
            plateau_lines.append(
                f"{lift}: fatigue plateau — e1RM stalled ({vals_str}) while RPE rising — {suggestion}"
            )
        else:
            suggestion = "consider a deload: drop top set by 10%, rebuild with 5×3 for 2 weeks"
            plateau_lines.append(
                f"{lift}: e1RM flat for 3 weeks ({vals_str}) — {suggestion}"
            )

    if plateau_lines:
        lines.append("")
        lines.append("--- Plateau Alerts ---")
        for pl in plateau_lines:
            lines.append(pl)

    # --- Training Gaps ---
    sets_by_muscle = vol.get("sets_by_muscle_group", {}) if vol else {}
    gap_groups = sorted(mg for mg, total in sets_by_muscle.items() if total == 0)
    if gap_groups:
        window_weeks = aw.get("weeks", 0)
        window_days = window_weeks * 7
        lines.append("")
        lines.append("--- Training Gaps ---")
        for mg in gap_groups:
            lines.append(f"No direct {mg.lower()} work in {window_days} days.")

    return "\n".join(lines)


def iso_week(date_str: str) -> str:
    """Return 'YYYY-Www' week label from ISO date string."""
    try:
        d = datetime.fromisoformat(date_str).date()
        return f"{d.isocalendar()[0]}-W{d.isocalendar()[1]:02d}"
    except ValueError:
        return "unknown"


def _build_adherence_line(adh: dict) -> str:
    """Build a human-readable adherence summary from score_adherence() output.

    Returns empty string when total == 0 (no scorable data).
    Example: 'Adherence: 6/9 hit target · 2 beat · 1 missed (Bench 70×5 → planned 72.5×5)'
    """
    if not adh or adh.get("total", 0) == 0:
        return ""

    total   = adh["total"]
    hit     = adh["hit"]
    beat    = adh["beat"]
    missed  = adh["missed"]

    missed_examples = [
        f"{d['exercise']} {d['actual']} → planned {d['planned']}"
        for d in adh.get("detail", [])
        if d.get("outcome") == "missed"
    ]
    example = f" ({missed_examples[0]})" if missed_examples else ""

    return (
        f"Adherence: {hit + beat}/{total} hit target "
        f"· {beat} beat · {missed} missed{example}"
    )


def format_for_trainer(entries: list[dict]) -> str:
    if not entries:
        return "No workout data available for this period."

    by_date: dict[str, dict] = {}
    for e in entries:
        d = e["date"]
        if d not in by_date:
            by_date[d] = {"session": e["session"], "exercises": []}
        by_date[d]["exercises"].append(e)

    total_sessions  = len(by_date)
    total_exercises = len(entries)

    lines = [
        f"**Period:** {min(by_date)} to {max(by_date)}",
        f"**Sessions:** {total_sessions}  |  **Exercise entries:** {total_exercises}",
        "",
    ]

    for d in sorted(by_date.keys()):
        session = by_date[d]["session"]
        try:
            day_label = datetime.fromisoformat(d).strftime("%a %d %b")
        except ValueError:
            day_label = d
        week = iso_week(d)
        lines.append(f"\n**{day_label}** ({week}) — {session}")

        for ex in by_date[d]["exercises"]:
            sets   = ex["sets"] if ex["sets"] is not None else "?"
            reps   = ex["reps"] if ex["reps"] is not None else "—"
            weight = ex["weight"] or "BW"
            top    = f"  ← top: {ex['top_set_kg']} kg" if ex["top_set_kg"] else ""
            lines.append(f"  - {ex['exercise']} [{ex['muscle_group']}]: {sets}×{reps} @ {weight}{top}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Fetch journal entries for work/life context
# ---------------------------------------------------------------------------

def fetch_journal_entries(weeks: int) -> list[dict]:
    """Fetch recent daily journal pages from Notion for work/life context."""
    if not NOTION_DATABASE_ID:
        return []

    since = (date.today() - timedelta(weeks=weeks)).isoformat()
    headers = {
        "Authorization": f"Bearer {NOTION_TOKEN}",
        "Content-Type": "application/json",
        "Notion-Version": NOTION_VERSION,
    }
    payload = {
        "filter": {"property": "Date", "date": {"on_or_after": since}},
        "sorts": [{"property": "Date", "direction": "ascending"}],
        "page_size": 50,
    }
    resp = requests.post(
        f"https://api.notion.com/v1/databases/{NOTION_DATABASE_ID}/query",
        headers=headers, json=payload, timeout=20,
    )
    resp.raise_for_status()
    pages = resp.json().get("results", [])

    entries = []
    for page in pages:
        props = page.get("properties", {})
        date_val    = (props.get("Date", {}).get("date") or {}).get("start", "")
        title_items = props.get("Name", {}).get("title", [])
        title       = title_items[0]["plain_text"] if title_items else ""

        try:
            blocks_resp = requests.get(
                f"https://api.notion.com/v1/blocks/{page['id']}/children?page_size=5",
                headers=headers, timeout=10,
            )
            blocks = blocks_resp.json().get("results", [])
            text_parts = []
            for b in blocks:
                btype = b.get("type", "")
                rich  = b.get(btype, {}).get("rich_text", [])
                text_parts.extend(r.get("plain_text", "") for r in rich)
            summary = " ".join(text_parts)[:500]
        except Exception:
            summary = ""

        entries.append({"date": date_val, "title": title, "summary": summary})

    log.info(f"Fetched {len(entries)} journal entries from last {weeks} weeks")
    return entries


# ---------------------------------------------------------------------------
# AI call
# ---------------------------------------------------------------------------

def call_ai(prompt: str) -> str:
    """Delegate to ai_client using the weekly report system prompt."""
    return ai_client.call_ai(
        prompt,
        WEEKLY_REPORT_PROMPT,
        label="Weekly report",
        max_tokens=2048,
    )


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _style_header(ws):
    """Bold + grey fill for the first row."""
    try:
        from openpyxl.styles import Font, PatternFill
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9D9D9")
    except Exception:
        pass


def export_to_excel(entries: list[dict], output_path: Path):
    """Export all workout data to Excel with 4 analytical sheets."""
    try:
        import openpyxl
    except ImportError:
        log.error("openpyxl not installed — run: pip install openpyxl --break-system-packages")
        return

    wb = openpyxl.Workbook()

    # ── Sheet 1: All Workouts ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All Workouts"
    ws1.append(["Date", "Session", "Exercise", "Muscle Group", "Sets", "Reps", "Weight", "Top Set (kg)"])
    for e in sorted(entries, key=lambda x: x["date"]):
        ws1.append([
            e["date"], e["session"], e["exercise"], e["muscle_group"],
            e["sets"], e["reps"], e["weight"], e["top_set_kg"],
        ])
    _style_header(ws1)

    # ── Sheet 2: Key Lifts Progress ──────────────────────────────────────────
    ws2 = wb.create_sheet("Key Lifts Progress")
    KEY_LIFTS = {
        "Bench Press": ["bench"],
        "Deadlift":    ["deadlift"],
        "Squat":       ["squat"],
        "OHP":         ["overhead press", "ohp"],
    }
    lift_data: dict[str, dict[str, float]] = {lift: {} for lift in KEY_LIFTS}
    for e in entries:
        if e["top_set_kg"] is None:
            continue
        name_lower = e["exercise"].lower()
        for lift, keywords in KEY_LIFTS.items():
            if any(kw in name_lower for kw in keywords):
                d = e["date"]
                if d not in lift_data[lift] or e["top_set_kg"] > lift_data[lift][d]:
                    lift_data[lift][d] = e["top_set_kg"]
                break

    all_dates = sorted({e["date"] for e in entries})
    ws2.append(["Date"] + list(KEY_LIFTS.keys()))
    for d in all_dates:
        ws2.append([d] + [lift_data[lift].get(d, "") for lift in KEY_LIFTS])
    _style_header(ws2)

    # ── Sheet 3: Volume by Week ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Volume by Week")
    volume: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    muscle_groups: set[str] = set()
    for e in entries:
        if e["date"] and e["sets"]:
            try:
                d    = datetime.fromisoformat(e["date"]).date()
                week = f"{d.isocalendar()[0]}-W{d.isocalendar()[1]:02d}"
                mg   = e["muscle_group"] or "Other"
                volume[week][mg] += e["sets"]
                muscle_groups.add(mg)
            except ValueError:
                pass

    mg_list = sorted(muscle_groups)
    ws3.append(["Week"] + mg_list)
    for week in sorted(volume.keys()):
        ws3.append([week] + [volume[week].get(mg, 0) for mg in mg_list])
    _style_header(ws3)

    # ── Sheet 4: Session Log ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Session Log")
    sessions_by_week: dict[str, dict] = defaultdict(
        lambda: {"Mon": 0, "Wed": 0, "Fri": 0, "Sat": 0, "Other": 0}
    )
    for e in entries:
        if e["date"]:
            try:
                d       = datetime.fromisoformat(e["date"]).date()
                week    = f"{d.isocalendar()[0]}-W{d.isocalendar()[1]:02d}"
                day_map = {0: "Mon", 2: "Wed", 4: "Fri", 5: "Sat"}
                day_key = day_map.get(d.weekday(), "Other")
                sessions_by_week[week][day_key] = 1
            except ValueError:
                pass

    ws4.append(["Week", "Mon (Chest)", "Wed (Deadlift)", "Fri (Squat)", "Sat (Arms)", "Total"])
    for week in sorted(sessions_by_week.keys()):
        s     = sessions_by_week[week]
        total = s["Mon"] + s["Wed"] + s["Fri"] + s["Sat"]
        ws4.append([week, s["Mon"], s["Wed"], s["Fri"], s["Sat"], total])
    _style_header(ws4)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info(f"Excel report saved: {output_path}")


# ---------------------------------------------------------------------------
# Post to Notion
# ---------------------------------------------------------------------------

def _parse_inline(text: str) -> list[dict]:
    """
    Convert inline markdown bold (**text**) into Notion rich_text segments.
    Returns a list of rich_text objects suitable for any Notion block.
    """
    parts = re.split(r'\*\*(.+?)\*\*', text)
    rich = []
    for i, part in enumerate(parts):
        if not part:
            continue
        is_bold = (i % 2 == 1)
        rich.append({
            "type": "text",
            "text": {"content": part[:2000]},
            **({"annotations": {"bold": True}} if is_bold else {}),
        })
    return rich or [{"type": "text", "text": {"content": ""}}]


def md_to_notion_blocks(md: str) -> list[dict]:
    """Convert markdown to Notion block list. Handles ##/###, - and * bullets, inline **bold**."""
    blocks = []
    for line in md.split("\n"):
        s = line.strip()
        if not s:
            continue
        if s.startswith("### "):
            blocks.append({
                "object": "block", "type": "heading_3",
                "heading_3": {"rich_text": _parse_inline(s[4:].strip())},
            })
        elif s.startswith("## "):
            blocks.append({
                "object": "block", "type": "heading_2",
                "heading_2": {"rich_text": _parse_inline(s[3:].strip())},
            })
        elif s.startswith("- ") or s.startswith("* "):
            content = s[2:].strip()
            blocks.append({
                "object": "block", "type": "bulleted_list_item",
                "bulleted_list_item": {"rich_text": _parse_inline(content)},
            })
        else:
            blocks.append({
                "object": "block", "type": "paragraph",
                "paragraph": {"rich_text": _parse_inline(s)},
            })
    return blocks


def post_to_notion(feedback_md: str):
    """Append a dated feedback block to the trainer Notion page."""
    if not NOTION_TRAINER_PAGE_ID:
        log.warning("NOTION_TRAINER_PAGE_ID not set — printing to stdout")
        print("\n" + "=" * 60 + "\n" + feedback_md + "\n" + "=" * 60)
        return

    headers = {
        "Authorization": f"Bearer {NOTION_TOKEN}",
        "Content-Type": "application/json",
        "Notion-Version": NOTION_VERSION,
    }

    today_label = date.today().strftime("%B %d, %Y")
    blocks = [
        {
            "object": "block", "type": "heading_2",
            "heading_2": {"rich_text": [{"type": "text", "text": {"content": f"Weekly Report — {today_label}"}}]},
        }
    ]
    blocks.extend(md_to_notion_blocks(feedback_md))
    blocks.append({"object": "block", "type": "divider", "divider": {}})

    resp = requests.patch(
        f"https://api.notion.com/v1/blocks/{NOTION_TRAINER_PAGE_ID}/children",
        headers=headers,
        json={"children": blocks},
        timeout=30,
    )

    if resp.status_code == 200:
        log.info(f"Weekly report posted to Notion page {NOTION_TRAINER_PAGE_ID}")
    else:
        log.error(f"Notion append failed {resp.status_code}: {resp.text[:300]}")
        print(feedback_md)


# ---------------------------------------------------------------------------
# Email delivery — structure copied from Morning Debrief sender.py
# ---------------------------------------------------------------------------

_FONT  = "-apple-system, BlinkMacSystemFont, 'Helvetica Neue', Helvetica, Arial, sans-serif"

# Debrief-matched colour palette (light theme)
_C_BG     = "#f2f2f1"   # outer background — warm light gray
_C_CARD   = "#ffffff"   # card background
_C_LABEL  = "#655e59"   # small-caps section labels
_C_BODY   = "#151514"   # primary text — near black
_C_BORDER = "#e1dfdd"   # divider lines
_C_FOOTER = "#97918d"   # footer text
_C_ROWBG  = "#f8f8f7"   # tinted block background


def _e(s) -> str:
    """HTML-escape a value."""
    from html import escape
    return escape(str(s) if s is not None else "")


# ---------------------------------------------------------------------------
# 1RM estimates + progression graph
# ---------------------------------------------------------------------------

def _extract_e1rm_estimates(metrics: dict) -> dict:
    """Pull current e1RM kg from compute_metrics() output for each benchmark lift."""
    result = {}
    for lift in ("bench", "deadlift", "squat"):
        data = metrics.get("strength_progression", {}).get(lift, {})
        if not data.get("insufficient_data") and data.get("current_e1rm"):
            result[lift] = data["current_e1rm"]
    return result


def _compute_e1rm_series(entries: list[dict]) -> dict:
    """Return {lift: [(iso_week_str, e1rm_kg), ...]} sorted chronologically."""
    if not _ANALYTICS_AVAILABLE:
        return {}

    weekly: dict = {"bench": {}, "deadlift": {}, "squat": {}}

    for e in entries:
        lift = identify_benchmark(e.get("exercise", ""))
        if lift is None:
            continue
        week = iso_week(e.get("date", ""))
        if week == "unknown":
            continue
        top_kg = e.get("top_set_kg")
        if not top_kg:
            continue
        top_kg = float(top_kg)
        if top_kg > 0 and (week not in weekly[lift] or top_kg > weekly[lift][week]):
            weekly[lift][week] = top_kg

    return {lift: sorted(wd.items()) for lift, wd in weekly.items()}


def _build_progression_svg(series: dict) -> str:
    """Inline SVG line chart of top set progression. Returns '' if insufficient data."""
    all_weeks = sorted({w for data in series.values() for w, _ in data})
    all_vals  = [v for data in series.values() for _, v in data]
    if len(all_weeks) < 2 or not all_vals:
        return ""

    W, H = 516, 190
    ML, MT, MR, MB = 44, 18, 96, 28
    pw = W - ML - MR
    ph = H - MT - MB

    n      = len(all_weeks)
    idx    = {w: i for i, w in enumerate(all_weeks)}
    y_min  = max(0.0, min(all_vals) - 10)
    y_max  = max(all_vals) + 12
    y_rng  = y_max - y_min

    def px(w: str) -> float:
        return ML + (idx[w] / (n - 1)) * pw

    def py(v: float) -> float:
        return MT + ph - ((v - y_min) / y_rng) * ph

    COLORS = {"bench": "#4A90D9", "deadlift": "#C94040", "squat": "#5BA858"}
    NAMES  = {"bench": "Bench",   "deadlift": "Deadlift", "squat": "Squat"}
    FONT   = "ui-sans-serif,system-ui,-apple-system,sans-serif"

    parts: list[str] = [
        f'<svg width="{W}" height="{H}" viewBox="0 0 {W} {H}" '
        f'xmlns="http://www.w3.org/2000/svg" style="display:block;max-width:100%;">'
    ]

    # Gridlines + y-axis labels (4 ticks)
    for i in range(4):
        v  = y_min + (y_rng / 3) * i
        yp = py(v)
        parts.append(
            f'<line x1="{ML}" y1="{yp:.1f}" x2="{ML+pw}" y2="{yp:.1f}" '
            f'stroke="#e1dfdd" stroke-width="1"/>'
        )
        parts.append(
            f'<text x="{ML-5}" y="{yp+4:.1f}" text-anchor="end" fill="#97918d" '
            f'font-size="10" font-family="{FONT}">{v:.0f}</text>'
        )

    # Lines + dots per lift
    final_points: list[tuple] = []
    for lift, data in series.items():
        if len(data) < 2:
            if len(data) == 1:
                final_points.append((lift, data[0][1]))
            continue
        color  = COLORS[lift]
        pts    = " ".join(f"{px(w):.1f},{py(v):.1f}" for w, v in data)
        parts.append(
            f'<polyline points="{pts}" fill="none" stroke="{color}" '
            f'stroke-width="2" stroke-linejoin="round" stroke-linecap="round"/>'
        )
        for w, v in data:
            parts.append(
                f'<circle cx="{px(w):.1f}" cy="{py(v):.1f}" r="2.5" fill="{color}"/>'
            )
        final_points.append((lift, data[-1][1]))

    # Right-side labels, sorted by final value descending so they render top→bottom
    final_points.sort(key=lambda t: -t[1])
    prev_y = None
    for lift, val in final_points:
        label_y = py(val) + 4
        if prev_y is not None and label_y < prev_y + 13:
            label_y = prev_y + 13
        prev_y = label_y
        parts.append(
            f'<text x="{ML+pw+7}" y="{label_y:.1f}" fill="{COLORS[lift]}" '
            f'font-size="11" font-weight="600" font-family="{FONT}">'
            f'{NAMES[lift]} {val:.0f}</text>'
        )

    # X-axis: first and last week labels
    for w, anchor in [(all_weeks[0], "start"), (all_weeks[-1], "end")]:
        parts.append(
            f'<text x="{px(w):.1f}" y="{MT+ph+18}" text-anchor="{anchor}" '
            f'fill="#97918d" font-size="10" font-family="{FONT}">{w}</text>'
        )

    parts.append("</svg>")
    return "\n".join(parts)


def _build_weight_svg(entries: list, e1rm_series: Optional[dict] = None) -> str:
    """Inline SVG line chart of bodyweight over time. Returns '' if fewer than 2 points.

    When e1rm_series is provided and has >=2 points in range, overlays a normalized
    e1RM trend line on a right Y-axis (orange, dashed).
    """
    if len(entries) < 2:
        return ""

    dates   = [e[0] for e in entries]
    weights = [e[1] for e in entries]

    # --- Try to build e1rm overlay ---
    overlay_points: list = []   # [(x_pos, y_pos), ...]
    e1rm_right_ticks = ""
    e1rm_label = ""
    e1rm_polyline = ""
    has_overlay = False

    try:
        if e1rm_series:
            # Pick first non-empty series: bench → deadlift → squat
            chosen_lift = None
            chosen_data = []
            for lift in ("bench", "deadlift", "squat"):
                data = e1rm_series.get(lift, [])
                if len(data) >= 2:
                    chosen_lift = lift
                    chosen_data = data
                    break

            if chosen_lift and chosen_data:
                bw_min_date = dates[0]
                bw_max_date = dates[-1]

                # Convert e1rm week strings to Monday ISO dates
                filtered = []
                for week_str, val in chosen_data:
                    mon_date = datetime.strptime(f"{week_str}-1", "%G-W%V-%u").date()
                    mon_str = mon_date.isoformat()
                    if bw_min_date <= mon_str <= bw_max_date:
                        filtered.append((mon_str, val))

                if len(filtered) >= 2:
                    has_overlay = True
                    W_overlay = 580  # wider to accommodate right axis
                else:
                    W_overlay = 560
            else:
                W_overlay = 560
        else:
            W_overlay = 560
    except Exception:
        has_overlay = False
        W_overlay = 560

    W, H, PAD = W_overlay, 140, 40
    min_w = min(weights) - 0.5
    max_w = max(weights) + 0.5
    span_w = max_w - min_w or 1.0

    def wx(i: int) -> float:
        return PAD + i * (W - 2 * PAD) / (len(weights) - 1)

    def wy(v: float) -> float:
        return H - PAD - (v - min_w) / span_w * (H - 2 * PAD)

    points = " ".join(f"{wx(i):.1f},{wy(w):.1f}" for i, w in enumerate(weights))
    dots   = "".join(
        f'<circle cx="{wx(i):.1f}" cy="{wy(w):.1f}" r="3" fill="#4A90E2"/>'
        for i, w in enumerate(weights)
    )

    delta     = weights[-1] - weights[0]
    direction = "▼" if delta < -0.1 else "▲" if delta > 0.1 else "→"
    current   = f"{weights[-1]:.1f} kg"
    delta_str = f"{direction} {abs(delta):.1f} kg" if abs(delta) > 0.1 else "→ stable"

    label_l = dates[0][5:]   # MM-DD
    label_r = dates[-1][5:]

    y_ticks = ""
    for v in [min_w + 0.5, (min_w + max_w) / 2, max_w - 0.5]:
        y_ticks += (
            f'<text x="{PAD - 4}" y="{wy(v):.1f}" text-anchor="end" '
            f'font-size="10" fill="#999">{v:.1f}</text>\n'
        )

    # Build overlay SVG elements if we have usable data
    if has_overlay:
        try:
            bw_min_date = dates[0]
            bw_max_date = dates[-1]

            # Re-derive filtered data (has_overlay guarantees this succeeds)
            chosen_data_full = []
            for lift in ("bench", "deadlift", "squat"):
                data = e1rm_series.get(lift, [])
                if len(data) >= 2:
                    chosen_data_full = data
                    break

            filtered = []
            for week_str, val in chosen_data_full:
                mon_date = datetime.strptime(f"{week_str}-1", "%G-W%V-%u").date()
                mon_str = mon_date.isoformat()
                if bw_min_date <= mon_str <= bw_max_date:
                    filtered.append((mon_str, val))

            e1rm_vals = [v for _, v in filtered]
            e1rm_min = min(e1rm_vals) - 5
            e1rm_max = max(e1rm_vals) + 5
            e1rm_span = e1rm_max - e1rm_min or 1.0

            def ey(v: float) -> float:
                return H - PAD - (v - e1rm_min) / e1rm_span * (H - 2 * PAD)

            # Date-based x position using linear interpolation
            from datetime import date as _date
            d_min = _date.fromisoformat(bw_min_date)
            d_max = _date.fromisoformat(bw_max_date)
            total_days = (d_max - d_min).days or 1

            def ex(date_str: str) -> float:
                d = _date.fromisoformat(date_str)
                ratio = (d - d_min).days / total_days
                return PAD + ratio * (W - 2 * PAD)

            overlay_pts = " ".join(
                f"{ex(ds):.1f},{ey(v):.1f}" for ds, v in filtered
            )

            # Right Y-axis tick labels at e1rm_min+5 and e1rm_max-5
            tick_lo = round(e1rm_min + 5)
            tick_hi = round(e1rm_max - 5)
            e1rm_right_ticks = (
                f'<text x="{W - 5}" y="{ey(tick_lo):.1f}" text-anchor="start" '
                f'font-size="10" fill="#E87040">{tick_lo}</text>\n'
                f'<text x="{W - 5}" y="{ey(tick_hi):.1f}" text-anchor="start" '
                f'font-size="10" fill="#E87040">{tick_hi}</text>\n'
            )

            e1rm_label = (
                f'<text x="{W - 35}" y="{PAD - 4}" text-anchor="start" '
                f'font-size="10" fill="#E87040">e1RM</text>\n'
            )

            e1rm_polyline = (
                f'<polyline points="{overlay_pts}" fill="none" stroke="#E87040" '
                f'stroke-width="1.5" stroke-dasharray="4 2"/>\n'
            )
        except Exception:
            # Fall back silently — render without overlay
            e1rm_right_ticks = ""
            e1rm_label = ""
            e1rm_polyline = ""

    return (
        f'<svg width="{W}" height="{H + 24}" xmlns="http://www.w3.org/2000/svg" '
        f'style="font-family:sans-serif;display:block;margin:12px auto">\n'
        f'  <text x="{W//2}" y="14" text-anchor="middle" font-size="13" '
        f'font-weight="bold" fill="#333">Bodyweight — {current} ({delta_str})</text>\n'
        f'  {y_ticks}'
        f'  {e1rm_right_ticks}'
        f'  {e1rm_label}'
        f'  <polyline points="{points}" fill="none" stroke="#4A90E2" stroke-width="2"/>\n'
        f'  {e1rm_polyline}'
        f'  {dots}\n'
        f'  <text x="{wx(0):.1f}" y="{H + 18}" text-anchor="middle" '
        f'font-size="10" fill="#888">{label_l}</text>\n'
        f'  <text x="{wx(len(weights)-1):.1f}" y="{H + 18}" text-anchor="middle" '
        f'font-size="10" fill="#888">{label_r}</text>\n'
        f'</svg>'
    )


def _build_muscle_pie_svg(sets_by_muscle: dict) -> str:
    """Inline SVG pie chart of muscle group split. Returns '' if no data."""
    total = sum(sets_by_muscle.values())
    if not sets_by_muscle or total == 0:
        return ""

    COLORS = [
        "#4A90E2", "#E87040", "#7CC576", "#9B59B6",
        "#F1C40F", "#E74C3C", "#1ABC9C", "#95A5A6",
    ]
    W, CX, CY, R = 560, 150, 120, 95

    slices, legend = "", ""
    angle = -math.pi / 2  # start at top
    sorted_items = sorted(sets_by_muscle.items(), key=lambda x: x[1], reverse=True)

    for idx, (muscle, sets) in enumerate(sorted_items):
        pct   = sets / total
        sweep = pct * 2 * math.pi
        x1    = CX + R * math.cos(angle)
        y1    = CY + R * math.sin(angle)
        angle += sweep
        x2    = CX + R * math.cos(angle)
        y2    = CY + R * math.sin(angle)
        large = 1 if sweep > math.pi else 0
        color = COLORS[idx % len(COLORS)]

        slices += (
            f'<path d="M{CX},{CY} L{x1:.2f},{y1:.2f} '
            f'A{R},{R} 0 {large},1 {x2:.2f},{y2:.2f} Z" '
            f'fill="{color}" stroke="white" stroke-width="1.5"/>\n'
        )

        lx = 270 + (idx // 5) * 160
        ly = 32 + (idx % 5) * 22
        legend += (
            f'<rect x="{lx}" y="{ly}" width="11" height="11" fill="{color}"/>'
            f'<text x="{lx + 15}" y="{ly + 10}" font-size="11" fill="#333">'
            f'{muscle} {pct * 100:.0f}%</text>\n'
        )

    return (
        f'<svg width="{W}" height="250" xmlns="http://www.w3.org/2000/svg" '
        f'style="font-family:sans-serif;display:block;margin:12px auto">\n'
        f'  <text x="{W//2}" y="16" text-anchor="middle" font-size="13" '
        f'font-weight="bold" fill="#333">Muscle Group Split — Last 7 Days</text>\n'
        f'  {slices}'
        f'  {legend}'
        f'</svg>'
    )


def _build_report_html(
    md: str,
    e1rm: Optional[dict] = None,
    progression_svg: str = "",
    muscle_svg: str = "",
    weight_svg: str = "",
) -> str:
    """
    Render the AI coaching report as HTML matching the Morning Debrief light theme.
    Table-based layout + inline CSS. Handles: ### / ## headings, **bold**, bullets, paragraphs.
    """
    today      = date.today()
    day_label  = today.strftime("%A, %B %-d, %Y")   # "Sunday, May 18, 2026"
    time_label = datetime.now().strftime("%H:%M")
    from_addr  = EMAIL_FROM or "debrief@michalgzyl.com"

    def boldify(t: str) -> str:
        return re.sub(
            r'\*\*(.+?)\*\*',
            lambda m: f'<strong style="color:{_C_BODY};font-weight:600;">' + _e(m.group(1)) + '</strong>',
            _e(t),
        )

    rows: list[str] = []
    in_ul = False

    for raw_line in md.split("\n"):
        s = raw_line.strip()
        is_bullet = s.startswith("- ") or (s.startswith("* ") and len(s) > 2)

        if not is_bullet and in_ul:
            rows.append('</ul></td></tr>')
            in_ul = False

        if not s:
            continue

        if s.startswith("### "):
            # Strip leading "1. " numbering — section label shown as small caps
            label = re.sub(r'^\d+\.\s*', '', s[4:].strip())
            rows.append(
                f'<tr><td style="padding:20px 0 6px;border-top:1px solid {_C_BORDER};">'
                f'<p style="margin:0;font-size:12px;font-weight:600;color:{_C_LABEL};'
                f'text-transform:uppercase;letter-spacing:0.09em;font-family:{_FONT};">'
                + _e(label) + '</p></td></tr>'
            )
        elif s.startswith("## "):
            rows.append(
                f'<tr><td style="padding:16px 0 4px;border-top:1px solid {_C_BORDER};">'
                f'<p style="margin:0;font-size:14px;font-weight:600;color:{_C_BODY};'
                f'font-family:{_FONT};">'
                + boldify(s[3:].strip()) + '</p></td></tr>'
            )
        elif is_bullet:
            if not in_ul:
                rows.append(
                    f'<tr><td style="padding:6px 0 0;">'
                    f'<ul style="margin:0;padding:0 0 0 18px;list-style:disc;">'
                )
                in_ul = True
            rows.append(
                f'<li style="margin:4px 0;font-size:14px;color:{_C_BODY};line-height:1.55;'
                f'font-family:{_FONT};">' + boldify(s[2:].strip()) + '</li>'
            )
        else:
            # "Next session: ..." lines get a tinted left-border block
            if s.lower().startswith("next session"):
                rows.append(
                    f'<tr><td style="padding:6px 0 2px;">'
                    f'<p style="margin:0;font-size:13px;color:{_C_BODY};line-height:1.55;'
                    f'background:{_C_ROWBG};border-left:3px solid {_C_BORDER};'
                    f'padding:6px 10px;border-radius:2px;font-family:{_FONT};">'
                    + boldify(s) + '</p></td></tr>'
                )
            else:
                rows.append(
                    f'<tr><td style="padding:3px 0;">'
                    f'<p style="margin:0;font-size:14px;color:{_C_BODY};line-height:1.6;'
                    f'font-family:{_FONT};">'
                    + boldify(s) + '</p></td></tr>'
                )

    if in_ul:
        rows.append('</ul></td></tr>')

    body_rows = "\n".join(rows)

    # --- 1RM estimates block ---
    e1rm_html = ""
    if e1rm:
        LIFT_LABELS = {"bench": "Bench", "deadlift": "Deadlift", "squat": "Squat"}
        cells = "".join(
            f'<td style="text-align:center;padding:0 16px;">'
            f'<p style="margin:0 0 2px;font-size:10px;font-weight:600;color:{_C_LABEL};'
            f'text-transform:uppercase;letter-spacing:0.07em;font-family:{_FONT};">'
            f'{LIFT_LABELS[lift]}</p>'
            f'<p style="margin:0;font-size:22px;font-weight:700;color:{_C_BODY};'
            f'letter-spacing:-0.3px;font-family:{_FONT};">{val:.1f}'
            f'<span style="font-size:13px;font-weight:400;color:{_C_LABEL};"> kg</span></p>'
            f'</td>'
            for lift, val in e1rm.items() if val
        )
        if cells:
            e1rm_html = (
                f'<tr><td style="padding:0 32px 16px;">'
                f'<table role="presentation" cellspacing="0" cellpadding="0" border="0" '
                f'style="width:100%;background:{_C_ROWBG};border-radius:8px;">'
                f'<tr><td style="padding:6px 0 4px;text-align:center;">'
                f'<p style="margin:0;font-size:10px;font-weight:600;color:{_C_LABEL};'
                f'text-transform:uppercase;letter-spacing:0.09em;font-family:{_FONT};">'
                f'Estimated 1RM (Epley)</p></td></tr>'
                f'<tr>{cells}</tr>'
                f'<tr><td colspan="3" style="padding:10px 0 4px;"></td></tr>'
                f'</table></td></tr>'
            )

    # --- Chart blocks (each with its own label) ---
    def _chart_block(label: str, content: str) -> str:
        return (
            f'<tr><td style="padding:0 32px 20px;">'
            f'<p style="margin:0 0 6px;font-size:10px;font-weight:600;color:{_C_LABEL};'
            f'text-transform:uppercase;letter-spacing:0.09em;font-family:{_FONT};">'
            f'{label}</p>'
            + content +
            f'</td></tr>'
        )

    svg_html = ""
    if progression_svg:
        svg_html += _chart_block("Top Set Progression", progression_svg)
    if muscle_svg:
        svg_html += _chart_block("Muscle Group Distribution — This Week", muscle_svg)
    if weight_svg:
        svg_html += _chart_block("Bodyweight Trend", weight_svg)

    return (
        '<!DOCTYPE html>\n'
        '<html lang="en">\n'
        '<head>\n'
        '<meta charset="utf-8">\n'
        '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
        '</head>\n'
        f'<body style="margin:0;padding:24px 12px;background:{_C_BG};font-family:{_FONT};">\n'
        '<table role="presentation" cellspacing="0" cellpadding="0" border="0" style="width:100%;">\n'
        '<tr><td align="center">\n'
        # Card
        f'<table role="presentation" cellspacing="0" cellpadding="0" border="0" '
        f'style="width:100%;max-width:580px;background:{_C_CARD};border-radius:12px;'
        f'overflow:hidden;border:1px solid {_C_BORDER};">\n'
        # Header
        f'<tr><td style="padding:28px 32px 20px;">'
        f'<p style="margin:0 0 8px;font-size:12px;font-weight:600;color:{_C_LABEL};'
        f'text-transform:uppercase;letter-spacing:0.09em;font-family:{_FONT};">'
        'WEEKLY COACHING REPORT</p>'
        f'<p style="margin:0;font-size:26px;font-weight:700;color:{_C_BODY};'
        f'letter-spacing:-0.3px;line-height:1.2;font-family:{_FONT};">'
        + _e(day_label) + '</p>'
        '</td></tr>\n'
        # 1RM estimates + progression graph
        + e1rm_html
        + svg_html
        # Body
        + f'<tr><td style="padding:4px 32px 28px;">\n'
        '<table role="presentation" cellspacing="0" cellpadding="0" border="0" style="width:100%;">\n'
        + body_rows + '\n'
        '</table>\n'
        '</td></tr>\n'
        # Footer
        f'<tr><td style="padding:14px 32px;border-top:1px solid {_C_BORDER};text-align:center;">'
        f'<p style="margin:0;color:{_C_FOOTER};font-size:11px;font-family:{_FONT};">'
        f'Generated {_e(time_label)} &nbsp;&middot;&nbsp; {_e(from_addr)}'
        '</p></td></tr>\n'
        '</table>\n'
        '</td></tr>\n'
        '</table>\n'
        '</body>\n'
        '</html>'
    )


def send_weekly_email(
    report_md: str,
    e1rm: Optional[dict] = None,
    progression_svg: str = "",
    muscle_svg: str = "",
    weight_svg: str = "",
    excel_path: Optional[Path] = None,
):
    """
    Send the weekly coaching report as HTML matching the Morning Debrief style.
    MIMEMultipart() → MIMEText(body, 'html') + optional Excel attachment.
    Envelope sender = SMTP_USER (authenticated identity); From header = EMAIL_FROM (alias).
    """
    if not all([SMTP_HOST, SMTP_USER, SMTP_PASSWORD, EMAIL_FROM, EMAIL_TO]):
        log.warning("SMTP not configured — skipping email (set SMTP_HOST/USER/PASSWORD/EMAIL_FROM/EMAIL_TO in .env)")
        return

    today = date.today()
    day_label = today.strftime("%A, %B %-d, %Y")   # "Sunday, May 18, 2026"
    subject = f"Weekly Coaching Report — {day_label}"

    html_body = _build_report_html(
        report_md, e1rm=e1rm,
        progression_svg=progression_svg, muscle_svg=muscle_svg, weight_svg=weight_svg,
    )

    # Identical structure to Morning Debrief sender.py
    msg = MIMEMultipart()
    msg["From"]       = EMAIL_FROM       # visible From (alias)
    msg["To"]         = EMAIL_TO
    msg["Subject"]    = subject
    msg["Date"]       = formatdate(localtime=True)
    msg["Message-ID"] = make_msgid(domain=EMAIL_FROM.split("@")[-1])
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    # Excel attachment (appended after the HTML body part)
    if excel_path and excel_path.exists():
        with open(excel_path, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{excel_path.name}"')
        msg.attach(part)
        log.info(f"Attaching Excel: {excel_path.name}")

    envelope_from = SMTP_USER
    envelope_to   = [a.strip() for a in EMAIL_TO.split(",") if a.strip()]
    context       = ssl.create_default_context()

    for attempt in (1, 2):
        try:
            if SMTP_PORT == 465:
                with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, context=context, timeout=30) as s:
                    s.login(SMTP_USER, SMTP_PASSWORD)
                    s.sendmail(envelope_from, envelope_to, msg.as_string())
            else:
                with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as s:
                    s.ehlo(); s.starttls(context=context); s.ehlo()
                    s.login(SMTP_USER, SMTP_PASSWORD)
                    s.sendmail(envelope_from, envelope_to, msg.as_string())
            log.info(f"Weekly report emailed to {EMAIL_TO} (attempt {attempt})")
            return
        except (smtplib.SMTPServerDisconnected, smtplib.SMTPConnectError,
                TimeoutError, ConnectionError) as exc:
            log.warning(f"SMTP transient failure (attempt {attempt}/2): {exc}")
            if attempt == 1:
                time.sleep(5)
        except Exception as exc:
            log.error(f"Email failed: {exc}")
            return


# ---------------------------------------------------------------------------
# Sleep/energy correlation helpers
# ---------------------------------------------------------------------------

def compute_sleep_workout_correlation(metrics_entries: list, workout_entries: list) -> str:
    """
    Return a correlation line if ≥3 bad-sleep days also have workout data.
    Compares average top-set weight on bad-sleep days vs the overall average.

    Returns empty string when not enough data or metrics DB not configured.
    """
    if not metrics_entries or not workout_entries:
        return ""

    # Build {date: "bad"|"ok"|"good"} from metrics
    sleep_by_date: dict = {}
    for m in metrics_entries:
        if m.get("sleep") and m.get("date"):
            sleep_by_date[m["date"]] = m["sleep"]

    bad_sleep_dates = {d for d, s in sleep_by_date.items() if s == "bad"}
    if not bad_sleep_dates:
        return ""

    # Build {date: [top_set_kg]} from workout entries
    tops_by_date: dict = {}
    for e in workout_entries:
        d = e.get("date", "")[:10]
        top = e.get("top_set_kg")
        if d and top is not None:
            tops_by_date.setdefault(d, []).append(float(top))

    # Bad-sleep days that also have workout data
    bad_with_workout = [d for d in bad_sleep_dates if d in tops_by_date]
    if len(bad_with_workout) < 3:
        return ""

    all_tops = [t for tops in tops_by_date.values() for t in tops]
    if not all_tops:
        return ""

    avg_all = sum(all_tops) / len(all_tops)
    bad_tops = [t for d in bad_with_workout for t in tops_by_date[d]]
    avg_bad = sum(bad_tops) / len(bad_tops)

    if avg_all == 0:
        return ""

    pct_diff = (avg_bad - avg_all) / avg_all * 100
    sign = "+" if pct_diff >= 0 else ""
    return f"Top sets on bad-sleep days: {sign}{pct_diff:.0f}% vs your average."


def build_sleep_energy_summary(metrics_entries: list, weeks: int) -> str:
    """Build a one-line sleep/energy snapshot for the weekly report prompt."""
    if not metrics_entries:
        return ""
    sleep_counts: dict = {}
    energy_counts: dict = {}
    for m in metrics_entries:
        if m.get("sleep"):
            sleep_counts[m["sleep"]] = sleep_counts.get(m["sleep"], 0) + 1
        if m.get("energy"):
            energy_counts[m["energy"]] = energy_counts.get(m["energy"], 0) + 1
    if not sleep_counts and not energy_counts:
        return ""
    parts = []
    if sleep_counts:
        parts.append("Sleep: " + " | ".join(f"{k} {v}d" for k, v in sorted(sleep_counts.items())))
    if energy_counts:
        parts.append("Energy: " + " | ".join(f"{k} {v}d" for k, v in sorted(energy_counts.items())))
    return f"\n\n## Sleep / Energy (last {weeks} weeks)\n" + " · ".join(parts)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Weekly Gym + Journal Report")
    parser.add_argument("--weeks",     type=int, default=4, help="Weeks to analyse (default: 4)")
    parser.add_argument("--dry-run",   action="store_true",  help="Print report, don't write to Notion")
    parser.add_argument("--preview",   action="store_true",  help="Render HTML to weekly-preview.html, don't send")
    parser.add_argument("--excel",     action="store_true",  help="Export workout history to Excel")
    parser.add_argument("--excel-all", action="store_true",  help="Export ALL history (52 weeks) to Excel")
    args = parser.parse_args()

    if not NOTION_TOKEN or not NOTION_WORKOUT_DB_ID:
        log.error("NOTION_TOKEN and NOTION_WORKOUT_DB_ID must be set in .env")
        sys.exit(1)

    log.info(f"Weekly Report — analysing last {args.weeks} weeks")

    # Fetch workout data
    pages = fetch_workout_entries(args.weeks)
    if not pages:
        log.warning("No workout data found.")
        return
    entries = [parse_entry(p) for p in pages]

    # Full-history Excel export (--excel-all fetches 52 weeks instead of the analysis window)
    if args.excel_all:
        all_pages   = fetch_workout_entries(52)
        all_entries = [parse_entry(p) for p in all_pages]
        export_to_excel(all_entries, EXCEL_PATH)
        log.info(f"Full-history Excel written: {EXCEL_PATH}")

    # Fetch journal entries for work/life context
    journal_entries = []
    if NOTION_DATABASE_ID:
        journal_entries = fetch_journal_entries(args.weeks)
    else:
        log.info("NOTION_DATABASE_ID not set — skipping journal context")

    # Build combined prompt + extract 1RM estimates
    e1rm_estimates: dict = {}
    metrics: dict = {}
    if _ANALYTICS_AVAILABLE:
        try:
            metrics = compute_metrics(entries, args.weeks)
            workout_section = format_metrics_for_llm(metrics)
            e1rm_estimates = _extract_e1rm_estimates(metrics)
        except Exception as exc:
            log.error(f"Analytics failed: {exc} — falling back to raw format")
            workout_section = format_for_trainer(entries)
    else:
        workout_section = format_for_trainer(entries)

    # Fetch all-time data for progression graph
    progression_svg = ""
    e1rm_series: dict = {}
    try:
        all_pages   = fetch_all_workout_entries()
        all_entries = [parse_entry(p) for p in all_pages]
        e1rm_series = _compute_e1rm_series(all_entries)
        progression_svg = _build_progression_svg(e1rm_series)
        log.info("Progression graph computed")
    except Exception as exc:
        log.warning(f"Progression graph failed (non-fatal): {exc}")

    # Bodyweight trend
    bw_entries: list = []
    weight_svg = ""
    try:
        bw_entries = fetch_bodyweight_entries(args.weeks)
        weight_svg = _build_weight_svg(bw_entries, e1rm_series if e1rm_series else None)
        log.info(f"Bodyweight trend: {len(bw_entries)} data point(s)")
    except Exception as exc:
        log.warning(f"Bodyweight trend failed (non-fatal): {exc}")

    # Muscle group pie chart — last 7 days only
    muscle_pie_svg = ""
    sets_7d: dict = {}
    try:
        cutoff_7d = (date.today() - timedelta(days=7)).isoformat()
        for e in entries:
            if e.get("date", "") >= cutoff_7d:
                mg = e.get("muscle_group", "")
                if mg:
                    sets_7d[mg] = sets_7d.get(mg, 0) + int(e.get("sets") or 0)
        muscle_pie_svg = _build_muscle_pie_svg(sets_7d)
        log.info(f"Muscle pie chart: {len(sets_7d)} group(s)")
    except Exception as exc:
        log.warning(f"Muscle pie chart failed (non-fatal): {exc}")

    # Bodyweight trend section for LLM
    bw_prompt_section = ""
    if bw_entries:
        current_bw = bw_entries[-1][1]
        delta_bw   = bw_entries[-1][1] - bw_entries[0][1]
        avg_bw     = sum(e[1] for e in bw_entries) / len(bw_entries)
        direction  = "losing" if delta_bw < -0.2 else "gaining" if delta_bw > 0.2 else "stable"
        bw_prompt_section = (
            f"\n\n## Bodyweight Trend ({len(bw_entries)} readings)\n"
            f"Current: {current_bw:.1f} kg | "
            f"Change over period: {delta_bw:+.1f} kg ({direction}) | "
            f"Average: {avg_bw:.1f} kg\n"
            f"Comment on the athlete's weight trend in the context of their training."
        )
        if (delta_bw < -0.2 and _ANALYTICS_AVAILABLE and metrics and
                any(not metrics.get("strength_progression", {}).get(lift, {}).get("insufficient_data")
                    and metrics["strength_progression"][lift].get("trend") in ("improving", "stable")
                    for lift in ("bench", "deadlift", "squat"))):
            bw_prompt_section += f"\nRecomposition signal: weight {delta_bw:+.1f} kg, strength held → recomposition on track."

    # Last-7-day muscle split section for LLM
    muscle_prompt_section = ""
    if sets_7d:
        total_7d  = sum(sets_7d.values())
        split_str = " | ".join(
            f"{mg} {s / total_7d * 100:.0f}%"
            for mg, s in sorted(sets_7d.items(), key=lambda x: -x[1])
        )
        muscle_prompt_section = (
            f"\n\n## Muscle Group Split — Last 7 Days\n{split_str}\n"
            f"Based on this split, suggest specific focus areas for the upcoming week's training."
        )

    # Sleep/energy metrics
    metrics_entries: list = []
    sleep_correlation_line = ""
    sleep_energy_section = ""
    try:
        metrics_entries = fetch_metrics_entries(args.weeks)
        if metrics_entries:
            log.info(f"Metrics: {len(metrics_entries)} daily metrics row(s)")
            sleep_correlation_line = compute_sleep_workout_correlation(metrics_entries, entries)
            sleep_energy_section   = build_sleep_energy_summary(metrics_entries, args.weeks)
            if sleep_correlation_line:
                log.info(f"Sleep correlation: {sleep_correlation_line}")
    except Exception as exc:
        log.warning(f"Metrics fetch failed (non-fatal): {exc}")

    # Plan adherence (alongside existing e1RM/PR/plateau trends — omit cleanly if no config or no data)
    adherence_line = ""
    if _ANALYTICS_AVAILABLE and _PLAN_CONFIG_AVAILABLE:
        try:
            plan_cfg = load_plan_config()
            if plan_cfg:
                seen_slots: set = set()
                all_slots: list = []
                for split_tpl in plan_cfg.get("templates", {}).values():
                    for slot_def in split_tpl:
                        slot_name = slot_def.get("slot", "")
                        if slot_name not in seen_slots:
                            seen_slots.add(slot_name)
                            all_slots.append(slot_def)
                adh = score_adherence(entries, args.weeks * 7, all_slots)
                adherence_line = _build_adherence_line(adh)
                if adherence_line:
                    log.info(f"Plan adherence: {adherence_line}")
                else:
                    log.info("Plan adherence: not enough scorable data")
            else:
                log.info("Plan adherence: no plan config — section omitted")
        except Exception as exc:
            log.warning(f"Plan adherence scoring failed (non-fatal): {exc}")

    workout_section += bw_prompt_section + muscle_prompt_section + sleep_energy_section
    if sleep_correlation_line:
        workout_section += f"\n\n{sleep_correlation_line}"
    if adherence_line:
        workout_section += f"\n\n## Plan Adherence\n{adherence_line}"

    if journal_entries:
        journal_section = f"\n\n## Daily Journal Entries (last {args.weeks} weeks)\n"
        for je in journal_entries:
            journal_section += f"\n**{je['date']}** — {je['title']}\n{je['summary']}\n"
        prompt = workout_section + journal_section
    else:
        prompt = workout_section

    log.info(f"Sending {len(entries)} workout entries + {len(journal_entries)} journal entries to AI")

    try:
        feedback = call_ai(prompt)
    except RuntimeError as e:
        log.error(str(e))
        sys.exit(1)

    if args.dry_run:
        print("\n" + "=" * 60)
        print(feedback)
        print("=" * 60)
        if e1rm_estimates:
            print("\n1RM estimates:", e1rm_estimates)
        print(f"Weight SVG: {len(weight_svg)} chars | Pie SVG: {len(muscle_pie_svg)} chars | Progression SVG: {len(progression_svg)} chars")
    elif args.preview:
        html_body = _build_report_html(
            feedback, e1rm=e1rm_estimates,
            progression_svg=progression_svg, muscle_svg=muscle_pie_svg, weight_svg=weight_svg,
        )
        preview_path = Path(__file__).parent / "weekly-preview.html"
        preview_path.write_text(html_body, encoding="utf-8")
        log.info("Preview written to %s", preview_path)
    else:
        post_to_notion(feedback)
        # Always generate/refresh Excel (so email always has an up-to-date attachment)
        export_to_excel(entries, EXCEL_PATH)
        send_weekly_email(
            feedback, e1rm=e1rm_estimates,
            progression_svg=progression_svg, muscle_svg=muscle_pie_svg, weight_svg=weight_svg,
            excel_path=EXCEL_PATH,
        )


if __name__ == "__main__":
    main()
